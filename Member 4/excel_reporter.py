# -*- coding: utf-8 -*-
"""
SBOM Excel Compliance Report Generator (Member 4 Self-Contained)
Generates a professionally styled, multi-sheet Excel workbook
based on 21 SBOM attributes from the client requirements.
"""

import json
import os
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------------
# Design System - Colours & Fonts
# ------------------------------------------------------------------
class Theme:
    # Primary palette
    PRIMARY_DARK   = "0D1B2A"   # Deep Navy
    PRIMARY_MID    = "1B2838"   # Dark Slate
    PRIMARY_ACCENT = "1F4E78"   # Steel Blue (header bg)
    ACCENT_GOLD    = "E8A838"   # Warm Gold accent
    WHITE          = "FFFFFF"
    LIGHT_BG       = "F4F6F9"   # Very light grey for alt rows

    # Criticality palette
    CRIT_CRITICAL_BG   = "FF4C4C"
    CRIT_CRITICAL_FONT = "FFFFFF"
    CRIT_HIGH_BG       = "FF9800"
    CRIT_HIGH_FONT     = "000000"
    CRIT_MEDIUM_BG     = "FFC107"
    CRIT_MEDIUM_FONT   = "000000"
    CRIT_LOW_BG        = "4CAF50"
    CRIT_LOW_FONT      = "FFFFFF"

    # Borders
    BORDER_COLOR = "B0BEC5"

    # Fonts
    @staticmethod
    def title_font():
        return Font(name="Segoe UI", size=18, bold=True, color=Theme.PRIMARY_DARK)

    @staticmethod
    def subtitle_font():
        return Font(name="Segoe UI", size=11, italic=True, color="546E7A")

    @staticmethod
    def header_font():
        return Font(name="Segoe UI", size=10, bold=True, color=Theme.WHITE)

    @staticmethod
    def data_font():
        return Font(name="Segoe UI", size=9, color="212121")

    @staticmethod
    def data_bold_font():
        return Font(name="Segoe UI", size=9, bold=True, color="212121")

    @staticmethod
    def kpi_value_font():
        return Font(name="Segoe UI", size=28, bold=True, color=Theme.PRIMARY_DARK)

    @staticmethod
    def kpi_label_font():
        return Font(name="Segoe UI", size=10, color="78909C")

    @staticmethod
    def header_fill():
        return PatternFill(start_color=Theme.PRIMARY_ACCENT, end_color=Theme.PRIMARY_ACCENT, fill_type="solid")

    @staticmethod
    def alt_row_fill():
        return PatternFill(start_color=Theme.LIGHT_BG, end_color=Theme.LIGHT_BG, fill_type="solid")

    @staticmethod
    def gold_fill():
        return PatternFill(start_color=Theme.ACCENT_GOLD, end_color=Theme.ACCENT_GOLD, fill_type="solid")

    @staticmethod
    def thin_border():
        s = Side(border_style="thin", color=Theme.BORDER_COLOR)
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def criticality_style(level):
        level = (level or "medium").lower()
        mapping = {
            "critical": (Theme.CRIT_CRITICAL_BG, Theme.CRIT_CRITICAL_FONT),
            "high":     (Theme.CRIT_HIGH_BG,     Theme.CRIT_HIGH_FONT),
            "medium":   (Theme.CRIT_MEDIUM_BG,   Theme.CRIT_MEDIUM_FONT),
            "low":      (Theme.CRIT_LOW_BG,      Theme.CRIT_LOW_FONT),
        }
        bg, fg = mapping.get(level, (Theme.CRIT_MEDIUM_BG, Theme.CRIT_MEDIUM_FONT))
        return (
            PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
            Font(name="Segoe UI", size=9, bold=True, color=fg)
        )


# ------------------------------------------------------------------
# Helper - Apply styles to a row
# ------------------------------------------------------------------
def style_header_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = Theme.header_fill()
        cell.font = Theme.header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Theme.thin_border()


def style_data_cell(cell, is_alt=False, center=False):
    cell.font = Theme.data_font()
    cell.border = Theme.thin_border()
    h = "center" if center else "left"
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
    if is_alt:
        cell.fill = Theme.alt_row_fill()


def auto_fit_columns(ws, min_width=10, max_width=45, start_row=1):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells[start_row - 1:]:
            val = str(cell.value or "")
            max_len = max(max_len, len(val) + 3)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


# ------------------------------------------------------------------
# Sheet 1 - Dashboard
# ------------------------------------------------------------------
def create_dashboard_sheet(wb, sbom):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "1F4E78"

    components = sbom.get("components", [])
    vulnerabilities = sbom.get("vulnerabilities", [])
    metadata = sbom.get("metadata", {})
    timestamp = metadata.get("timestamp", "N/A")
    spec = sbom.get("specVersion", "N/A")

    # Count criticalities
    crit_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    for comp in components:
        props = {p["name"]: p["value"] for p in comp.get("properties", [])}
        lvl = props.get("criticality", "medium").lower()
        if lvl in crit_counts:
            crit_counts[lvl] += 1

    # Count licenses
    license_set = set()
    for comp in components:
        for lic in comp.get("licenses", []):
            license_set.add(lic.get("license", {}).get("name", "Unknown"))

    # -- Title block --
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="SBOM Compliance Dashboard")
    title_cell.font = Theme.title_font()
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    sub_text = "CycloneDX v{} | Generated: {} | Report Date: {}".format(
        spec, timestamp, datetime.datetime.now().strftime('%Y-%m-%d %H:%M'))
    sub_cell = ws.cell(row=2, column=1, value=sub_text)
    sub_cell.font = Theme.subtitle_font()
    ws.row_dimensions[2].height = 22

    # -- KPI Cards (row 4-6) --
    kpis = [
        ("Total Components", str(len(components))),
        ("Known Vulns", str(len(vulnerabilities))),
        ("Unique Licenses", str(len(license_set))),
        ("Critical", str(crit_counts["critical"])),
        ("High", str(crit_counts["high"])),
        ("Medium", str(crit_counts["medium"])),
        ("Low", str(crit_counts["low"])),
    ]

    card_fills = [
        PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
        PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
        PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid"),
        PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"),
    ]

    row_start = 4
    for i, (label, value) in enumerate(kpis):
        col = i + 1
        # Value cell
        v_cell = ws.cell(row=row_start, column=col, value=value)
        v_cell.font = Theme.kpi_value_font()
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        v_cell.fill = card_fills[i % len(card_fills)]
        v_cell.border = Theme.thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 22
        ws.row_dimensions[row_start].height = 50

        # Label cell
        l_cell = ws.cell(row=row_start + 1, column=col, value=label)
        l_cell.font = Theme.kpi_label_font()
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        l_cell.fill = card_fills[i % len(card_fills)]
        l_cell.border = Theme.thin_border()
    ws.row_dimensions[row_start + 1].height = 25

    # -- Component summary table (row 8+) --
    summary_row = 8
    ws.cell(row=summary_row - 1, column=1, value="Component Overview").font = Font(
        name="Segoe UI", size=13, bold=True, color=Theme.PRIMARY_DARK)
    summary_headers = ["#", "Component", "Version", "License", "Criticality", "Vulnerabilities", "PURL"]
    for ci, h in enumerate(summary_headers, 1):
        ws.cell(row=summary_row, column=ci, value=h)
    style_header_row(ws, summary_row, len(summary_headers))
    ws.row_dimensions[summary_row].height = 26

    vuln_map = {}
    for v in vulnerabilities:
        for a in v.get("affects", []):
            ref = a.get("ref", "")
            vuln_map.setdefault(ref, []).append(v.get("id", ""))

    for idx, comp in enumerate(components, 1):
        r = summary_row + idx
        is_alt = idx % 2 == 0
        bom_ref = comp.get("bom-ref", comp.get("purl", ""))
        props = {p["name"]: p["value"] for p in comp.get("properties", [])}
        crit = props.get("criticality", "medium")
        lics = ", ".join([l.get("license", {}).get("name", "") for l in comp.get("licenses", [])])
        vulns = ", ".join(vuln_map.get(bom_ref, [])) or "None"

        vals = [idx, comp.get("name", ""), comp.get("version", ""), lics,
                crit.upper(), vulns, comp.get("purl", "")]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            style_data_cell(cell, is_alt=is_alt, center=(ci in [1, 3, 5]))
            if ci == 5:  # criticality
                fill, font = Theme.criticality_style(crit)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 20

    auto_fit_columns(ws, start_row=summary_row)


# ------------------------------------------------------------------
# Sheet 2 - Full 21-Attribute Component Data
# ------------------------------------------------------------------
def create_component_data_sheet(wb, sbom):
    ws = wb.create_sheet("Component Data (21 Attr)")
    ws.sheet_properties.tabColor = "E8A838"

    metadata = sbom.get("metadata", {})
    author = (metadata.get("authors") or [{}])[0].get("name", "Unknown")
    timestamp = metadata.get("timestamp", "Unknown")
    components = sbom.get("components", [])
    vulnerabilities = sbom.get("vulnerabilities", [])
    deps = sbom.get("dependencies", [])
    dep_map = {d.get("ref"): d.get("dependsOn", []) for d in deps}

    vuln_map = {}
    for v in vulnerabilities:
        for a in v.get("affects", []):
            vuln_map.setdefault(a.get("ref", ""), []).append(v.get("id", ""))

<<<<<<< HEAD
    # Title - merged up to AB for 28 columns
    ws.merge_cells("A1:AB1")
    ws.cell(row=1, column=1, value="SBOM Component Data - All 21 Client Attributes").font = Theme.title_font()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:AB2")
=======
    # Title
    ws.merge_cells("A1:AC1")
    ws.cell(row=1, column=1, value="SBOM Component Data - All 21 Client Attributes + Enterprise Enhancements").font = Theme.title_font()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:AC2")
>>>>>>> 65ff5e2d424ac770651a791555a68d806c5799b0
    ws.cell(row=2, column=1, value="Author: {}  |  Generated: {}".format(author, timestamp)).font = Theme.subtitle_font()

    # Headers - exact 21 attribute names from client requirements + sub-attributes & audit proof
    headers = [
        "Sr No",                     # index
        "Component Name",            # 1
        "Component Version",         # 2
        "Component Description",     # 3
        "Component Supplier",        # 4
        "Component License",         # 5
        "Component Origin",          # 6
        "Component Dependencies",    # 7
        "Vulnerabilities",           # 8
        "Patch Status",              # 9
        "Release Date",              # 10
        "End-of-Life (EOL) Date",    # 11
        "Criticality",               # 12
        "Criticality Rationale",     # (sub-12)
        "Usage Restrictions",        # 13
        "Checksums or Hashes",       # 14
        "Comments or Notes",         # 15
        "Author of SBOM Data",       # 16
        "Timestamp",                 # 17
        "Executable Property",       # 18
        "Executable Evidence",       # (sub-18)
        "Archive Property",          # 19
        "Archive Metadata Details",  # (sub-19)
        "Structured Property",       # 20
        "Unique Identifier (PURL)",  # 21
        "Trust Score",               # (Score)
        "Identification Evidence",   # (Reasons)
        "Repository Source",         # (Ecosystem Registry)
    ]

    header_row = 4
    for ci, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=ci, value=h)
    style_header_row(ws, header_row, len(headers))
    ws.row_dimensions[header_row].height = 30

    # Column alignment rules: center these column indices (1-based)
    center_cols = {1, 3, 11, 12, 13, 15, 19, 20, 22, 24, 26}

    for idx, comp in enumerate(components, 1):
        r = header_row + idx
        is_alt = idx % 2 == 0

        props = {p["name"]: p["value"] for p in comp.get("properties", [])}
        bom_ref = comp.get("bom-ref") or comp.get("purl", "")

        # Resolve dependencies
        child_refs = dep_map.get(bom_ref, [])
        dep_names = []
        for ref in child_refs:
            dep_names.append(ref.split("/")[-1].split("@")[0] if "/" in ref else ref)
        dep_str = ", ".join(dep_names) if dep_names else "None"

        # Resolve vulnerabilities
        comp_vulns = vuln_map.get(bom_ref, [])
        vuln_str = ", ".join(comp_vulns) if comp_vulns else "None"

        # Resolve hashes
        hashes = comp.get("hashes", [])
        hash_str = ", ".join(["{}:{}...".format(h["alg"], h["content"][:16]) for h in hashes]) if hashes else "N/A"

        # Resolve license
        lic_str = ", ".join([l.get("license", {}).get("name", "") for l in comp.get("licenses", [])]) or "Unknown"

        crit = props.get("criticality", "medium")

        row_values = [
            idx,
            comp.get("name", ""),
            comp.get("version", ""),
            comp.get("description", ""),
            comp.get("supplier", {}).get("name", ""),
            lic_str,
            props.get("origin", "open-source"),
            dep_str,
            vuln_str,
            props.get("patch_status", "up-to-date"),
            props.get("release_date", ""),
            props.get("eol_date", ""),
            crit.upper(),
            props.get("criticality_reason", "None"),
            props.get("usage_restrictions", "None"),
            hash_str,
            props.get("comments", ""),
            author,
            timestamp,
            props.get("executable", "false"),
            props.get("executable_evidence", "None"),
            props.get("archive", "true"),
            props.get("archive_metadata", "None"),
            props.get("structured", "CycloneDX JSON"),
            comp.get("purl", ""),
            props.get("trust_score", "0%"),
            props.get("evidence_findings", "None"),
            props.get("repository_source", "Unknown"),
        ]

        for ci, val in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            style_data_cell(cell, is_alt=is_alt, center=(ci in center_cols))

            # Special criticality column styling (now column 13)
            if ci == 13:
                fill, font = Theme.criticality_style(crit)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Bold component name
            if ci == 2:
                cell.font = Theme.data_bold_font()

        ws.row_dimensions[r].height = 22

    auto_fit_columns(ws, start_row=header_row)



# ------------------------------------------------------------------
# Sheet 3 - Vulnerability Matrix
# ------------------------------------------------------------------
def create_vulnerability_sheet(wb, sbom):
    ws = wb.create_sheet("Vulnerability Matrix")
    ws.sheet_properties.tabColor = "FF4C4C"

    vulnerabilities = sbom.get("vulnerabilities", [])

    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="Vulnerability Matrix").font = Theme.title_font()
    ws.row_dimensions[1].height = 35

    headers = ["#", "CVE / Vuln ID", "Severity", "Source", "Description",
               "Affected Component", "Status", "Detail URL"]
    header_row = 3
    for ci, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=ci, value=h)
    style_header_row(ws, header_row, len(headers))
    ws.row_dimensions[header_row].height = 28

    severity_fills = {
        "critical": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "high":     PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "medium":   PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        "low":      PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        "unknown":  PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    }

    if not vulnerabilities:
        ws.merge_cells("A4:H4")
        cell = ws.cell(row=4, column=1, value="[OK] No known vulnerabilities found in this SBOM scan.")
        cell.font = Font(name="Segoe UI", size=12, bold=True, color="2E7D32")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 30
    else:
        for idx, vuln in enumerate(vulnerabilities, 1):
            r = header_row + idx
            is_alt = idx % 2 == 0

            vid = vuln.get("id", "")
            ratings = vuln.get("ratings", [])
            severity = ratings[0].get("severity", "unknown") if ratings else "unknown"
            source_obj = vuln.get("source", {})
            source = source_obj.get("name", "") if isinstance(source_obj, dict) else str(source_obj)
            desc = vuln.get("description", "")
            affects = vuln.get("affects", [])
            affected_parts = []
            for a in affects:
                ref = a.get("ref", "")
                affected_parts.append(ref.split("/")[-1].split("@")[0] if "/" in ref else ref)
            affected = ", ".join(affected_parts) if affected_parts else "N/A"

            # Get analysis status
            analysis = vuln.get("analysis", {})
            status = analysis.get("state", "unknown") if isinstance(analysis, dict) else "unknown"

            detail = source_obj.get("url", "") if isinstance(source_obj, dict) else ""

            vals = [idx, vid, severity.upper(), source, desc[:120], affected, status, detail]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                style_data_cell(cell, is_alt=is_alt, center=(ci in [1, 3, 7]))
                if ci == 3:
                    sev_fill = severity_fills.get(severity.lower(), severity_fills["unknown"])
                    cell.fill = sev_fill
                    cell.font = Theme.data_bold_font()
            ws.row_dimensions[r].height = 22

    auto_fit_columns(ws, start_row=header_row)


# ------------------------------------------------------------------
# Sheet 4 - Legend & Info
# ------------------------------------------------------------------
def create_legend_sheet(wb):
    ws = wb.create_sheet("Legend & Info")
    ws.sheet_properties.tabColor = "4CAF50"

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="Report Legend & Attribute Definitions").font = Theme.title_font()
    ws.row_dimensions[1].height = 35

    # Criticality legend
    ws.cell(row=3, column=1, value="Criticality Levels").font = Font(
        name="Segoe UI", size=12, bold=True, color=Theme.PRIMARY_DARK)

    legend = [
        ("CRITICAL", Theme.CRIT_CRITICAL_BG, Theme.CRIT_CRITICAL_FONT,
         "Core component; failure causes system-wide outage or critical security breach."),
        ("HIGH", Theme.CRIT_HIGH_BG, Theme.CRIT_HIGH_FONT,
         "Important component; failure significantly impacts functionality or security."),
        ("MEDIUM", Theme.CRIT_MEDIUM_BG, Theme.CRIT_MEDIUM_FONT,
         "Standard component; moderate impact on operations."),
        ("LOW", Theme.CRIT_LOW_BG, Theme.CRIT_LOW_FONT,
         "Minor or utility component; minimal operational impact."),
    ]

    for i, (label, bg, fg, desc) in enumerate(legend):
        r = 4 + i
        cell_label = ws.cell(row=r, column=1, value=label)
        cell_label.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell_label.font = Font(name="Segoe UI", size=10, bold=True, color=fg)
        cell_label.alignment = Alignment(horizontal="center", vertical="center")
        cell_label.border = Theme.thin_border()

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell_desc = ws.cell(row=r, column=2, value=desc)
        cell_desc.font = Theme.data_font()
        cell_desc.border = Theme.thin_border()
        cell_desc.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 24

    # 21 Attribute definitions
    attr_start = 10
    ws.cell(row=attr_start, column=1, value="21 Client-Required SBOM Attributes").font = Font(
        name="Segoe UI", size=12, bold=True, color=Theme.PRIMARY_DARK)

    attr_headers = ["Sr No", "Attribute Title", "Attribute Description"]
    for ci, h in enumerate(attr_headers, 1):
        ws.cell(row=attr_start + 1, column=ci, value=h)
    style_header_row(ws, attr_start + 1, len(attr_headers))

    attributes = [
        (1, "Component Name", "The name of the software component or library included in the SBOM."),
        (2, "Component Version", "The version number or identifier of the software component."),
        (3, "Component Description", "A brief description of the functionality and purpose of the component."),
        (4, "Component Supplier", "The entity that supplies the software component (vendor, OSS project, etc.)."),
        (5, "Component License", "The license under which the component is distributed."),
        (6, "Component Origin", "Source classification: proprietary, open-source, or third-party."),
        (7, "Component Dependencies", "Other components this component depends on, including names and versions."),
        (8, "Vulnerabilities", "Known CVEs or security weaknesses associated with the component."),
        (9, "Patch Status", "Whether patches or updates are available for known vulnerabilities."),
        (10, "Release Date", "The date when the component version was released."),
        (11, "End-of-Life (EOL) Date", "The date when support/maintenance is scheduled to end."),
        (12, "Criticality", "Importance to overall functionality or security (critical/high/medium/low)."),
        (13, "Usage Restrictions", "Export control restrictions, IP rights, or other usage limitations."),
        (14, "Checksums or Hashes", "Cryptographic hashes to verify integrity and authenticity."),
        (15, "Comments or Notes", "Additional annotations relevant to the component."),
        (16, "Author of SBOM Data", "The entity that created the SBOM data for this component."),
        (17, "Timestamp", "Date and time of SBOM data assembly."),
        (18, "Executable Property", "Whether the component can be executed as a standalone binary."),
        (19, "Archive Property", "Whether the component is stored as an archive or compressed file."),
        (20, "Structured Property", "The organized data format of the component listing (e.g., CycloneDX JSON)."),
        (21, "Unique Identifier", "A distinct code (PURL) assigned to each software component."),
    ]

    for i, (num, title, desc) in enumerate(attributes):
        r = attr_start + 2 + i
        is_alt = i % 2 == 0
        vals = [num, title, desc]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            style_data_cell(cell, is_alt=is_alt, center=(ci == 1))
        ws.row_dimensions[r].height = 22

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 50


# ------------------------------------------------------------------
# Main Entry Point
# ------------------------------------------------------------------
def generate_excel_report(enriched_sbom_path, output_excel_path):
    """Generate the full multi-sheet Excel compliance report."""
    print("[*] Reading enriched SBOM from: {}".format(enriched_sbom_path))
    if not os.path.exists(enriched_sbom_path):
        print("[!] Error: File not found -> {}".format(enriched_sbom_path))
        return False

    with open(enriched_sbom_path, "r", encoding="utf-8") as f:
        sbom = json.load(f)

    wb = openpyxl.Workbook()

    print("[*] Building Dashboard sheet...")
    create_dashboard_sheet(wb, sbom)

    print("[*] Building Component Data (21 Attributes) sheet...")
    create_component_data_sheet(wb, sbom)

    print("[*] Building Vulnerability Matrix sheet...")
    create_vulnerability_sheet(wb, sbom)

    print("[*] Building Legend & Info sheet...")
    create_legend_sheet(wb)

    # Freeze panes for data sheets
    for sheet_name in ["Component Data (21 Attr)", "Vulnerability Matrix"]:
        sheet = wb[sheet_name]
        if sheet_name == "Component Data (21 Attr)":
            sheet.freeze_panes = "A5"
        elif sheet_name == "Vulnerability Matrix":
            sheet.freeze_panes = "A4"

    os.makedirs(os.path.dirname(os.path.abspath(output_excel_path)), exist_ok=True)
    try:
        wb.save(output_excel_path)
    except PermissionError:
        print("\n[!] Error: Permission denied when saving to '{}'.".format(output_excel_path))
        print("    Please close the Excel file if it is currently open in Microsoft Excel or another reader, and try again.\n")
        return False
    except Exception as e:
        print("\n[!] Error saving Excel workbook: {}\n".format(e))
        return False

    print("[+] Excel report saved -> {}".format(output_excel_path))
    print("    Sheets: Dashboard | Component Data (21 Attr) | Vulnerability Matrix | Legend & Info")
    return True


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    enriched_path = os.path.join(base_dir, "sbom_final.json")
    output_path = os.path.join(base_dir, "sbom_report.xlsx")

    if len(sys.argv) > 1:
        enriched_path = sys.argv[1]
    if len(sys.argv) > 2:
        output_path = sys.argv[2]

    success = generate_excel_report(enriched_path, output_path)
    sys.exit(0 if success else 1)
